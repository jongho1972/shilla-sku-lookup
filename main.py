"""신라인터넷면세점 SKU 조회 웹앱.

SKU 번호 입력 → 국문/영문 브랜드명, 상품명, REF.NO, 상품문의 전화번호 조회.
신라 API: ajaxProducts (CSRF 토큰 필요) + 상품 상세 페이지 파싱.
"""

import asyncio
import json
import logging
import re
from io import BytesIO
from pathlib import Path
from typing import List
from urllib.parse import quote

from bs4 import BeautifulSoup
from curl_cffi import requests as creq
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, Response
from pydantic import BaseModel

logger = logging.getLogger(__name__)

app = FastAPI(title="신라 SKU 조회")

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
# 사이트별 엔드포인트 — kr: 국문몰, cn: 중문몰, en: 영문몰, jp: 일문몰 (동일 API 구조)
SITE_CONF = {
    "kr": {
        "search": "https://m.shilladfs.com/estore/kr/ko/search?query={q}",
        "ajax": "https://m.shilladfs.com/estore/kr/ko/ajaxProducts",
        "detail_m": "https://m.shilladfs.com/estore/kr/ko/p/{code}",   # 스크래핑용
        "detail_pc": "https://www.shilladfs.com/estore/kr/ko/p/{code}",  # 링크 노출용
        "accept_language": "ko-KR,ko;q=0.9",
        "parse_detail": True,   # 상세 페이지에서 전화번호 등 보조 파싱
    },
    "cn": {
        "search": "https://m.shilladutyfree.cn/estore/kr/zh/search?query={q}",
        "ajax": "https://m.shilladutyfree.cn/estore/kr/zh/ajaxProducts",
        "detail_m": "https://m.shilladutyfree.cn/estore/kr/zh/p/{code}",
        "detail_pc": "https://www.shilladutyfree.cn/estore/kr/zh/p/{code}",
        "accept_language": "zh-CN,zh;q=0.9",
        "parse_detail": False,  # 중문·영문·일문몰 상세 페이지에는 상품문의 전화번호 구조 없음
    },
    "en": {
        "search": "https://m.shilladfs.com/estore/kr/en/search?query={q}",
        "ajax": "https://m.shilladfs.com/estore/kr/en/ajaxProducts",
        "detail_m": "https://m.shilladfs.com/estore/kr/en/p/{code}",
        "detail_pc": "https://www.shilladfs.com/estore/kr/en/p/{code}",
        "accept_language": "en-US,en;q=0.9",
        "parse_detail": False,
    },
    "jp": {
        "search": "https://m.shilladfs.com/estore/kr/ja/search?query={q}",
        "ajax": "https://m.shilladfs.com/estore/kr/ja/ajaxProducts",
        "detail_m": "https://m.shilladfs.com/estore/kr/ja/p/{code}",
        "detail_pc": "https://www.shilladfs.com/estore/kr/ja/p/{code}",
        "accept_language": "ja-JP,ja;q=0.9",
        "parse_detail": False,
    },
}

_PHONE_RE = re.compile(r"\d{2,3}-\d{3,4}-\d{4}")
_CSRF_RE = re.compile(r"CSRFToken['\"\s:=]+([0-9a-f-]{36})")


def _site_conf(site: str) -> dict:
    conf = SITE_CONF.get(site)
    if not conf:
        raise HTTPException(status_code=400, detail="site는 kr·cn·en·jp만 지원합니다")
    return conf


def _make_session(conf: dict) -> creq.Session:
    sess = creq.Session(impersonate="chrome")
    sess.headers.update({"User-Agent": UA, "Accept-Language": conf["accept_language"]})
    return sess


def _get_csrf(sess: creq.Session, conf: dict, query: str) -> str:
    page = sess.get(conf["search"].format(q=quote(query)), timeout=15)
    m = _CSRF_RE.search(page.text)
    return m.group(1) if m else ""


def _extract_category(hit: dict) -> str:
    """검색 API 응답의 "코드:카테고리명" 목록에서 카테고리명 추출 (2뎁스가 더 구체적이면 우선)."""
    for cat_key in ("disp2DepthCategoryList", "disp1DepthCategoryList"):
        cat_items = hit.get(cat_key)
        if isinstance(cat_items, list) and cat_items:
            last = str(cat_items[-1])
            category = last.split(":", 1)[-1].strip() if ":" in last else last.strip()
            if category:
                return category
    return ""


def _fetch_product(sku: str, site: str = "kr") -> dict:
    conf = SITE_CONF[site]
    sess = _make_session(conf)

    # 1. 검색 페이지에서 CSRF 토큰 획득
    token = _get_csrf(sess, conf, sku)

    # 2. SKU로 상품 검색
    body = {
        "json": json.dumps({
            "category": "", "size": "10", "page": 0,
            "text": sku, "within": "", "query": sku,
            "pagination": "", "condition": {"discountRate": "0"},
        }, ensure_ascii=False),
        "CSRFToken": token,
    }
    r = sess.post(conf["ajax"], data=body,
                  headers={"X-Requested-With": "XMLHttpRequest"}, timeout=15)
    r.raise_for_status()
    results = r.json().get("results", [])

    if not results:
        return {}

    # skuNo 필드로 정확 매칭, 없으면 미조회 처리
    hit = next((it for it in results if it.get("skuNo") == sku), None)
    if not hit:
        return {}

    code = hit.get("code", "")
    brand_cat = hit.get("brandCategory") or {}
    brand_kr = (hit.get("brandName") or brand_cat.get("brandName") or "").strip()
    brand_en = (brand_cat.get("enName") or hit.get("brandEnName") or "").strip()

    category = _extract_category(hit)

    product_name = hit.get("productNameForDisp") or hit.get("name") or ""

    # 3. 상세 페이지에서 영문 브랜드명, 전화번호 파싱 (REF.NO는 API 응답에서 직접 사용)
    #    - strong.info_brand: "한글명 | 영문명"
    #    - strong.number_title → 부모 li 텍스트: 상품 문의
    ref_no = hit.get("refNo", "") or code
    phone = ""
    if code and conf["parse_detail"]:
        try:
            dr = sess.get(conf["detail_m"].format(code=code), timeout=15)
            soup = BeautifulSoup(dr.text, "html.parser")

            # info_brand 로 영문명 보완 (brandCategory.enName 없는 경우 폴백)
            if not brand_en:
                info_brand = soup.select_one("strong.info_brand")
                if info_brand:
                    ib_text = info_brand.get_text(strip=True)
                    if " | " in ib_text:
                        brand_kr, brand_en = [b.strip() for b in ib_text.split(" | ", 1)]

            # 상품유형: API 응답에서 못 얻었을 때만 브레드크럼 마지막 활성 항목으로 보조
            if not category:
                bc_items = soup.select("ul.breadcrumb_box li.on")
                if bc_items:
                    category = bc_items[-1].get_text(strip=True)

            # 상품 문의 전화번호 파싱
            for s in soup.select("strong.number_title"):
                label = s.get_text(strip=True)
                li = s.parent
                value = li.get_text(strip=True).replace(label, "", 1).strip()
                if "REF" in label.upper() and not hit.get("refNo"):
                    ref_no = value or ref_no
                elif "문의" in label:
                    m2 = _PHONE_RE.search(value)
                    if m2:
                        phone = m2.group(0)

        except Exception as e:
            logger.warning("상세 페이지 파싱 실패 code=%s: %s", code, e)

    return {
        "sku": sku,
        "ref_no": ref_no,
        "brand_kr": brand_kr,
        "brand_en": brand_en,
        "category": category,
        "product_name": product_name,
        "phone": phone,
        "detail_url": conf["detail_pc"].format(code=code) if code else "",
    }


@app.get("/shilla_logo.png")
async def logo():
    return FileResponse(Path(__file__).parent / "shilla_logo.png", media_type="image/png")


@app.get("/healthz")
async def healthz():
    return {"ok": True}


@app.get("/", response_class=HTMLResponse)
async def root():
    return FileResponse(Path(__file__).parent / "index.html")


def _item_from_hit(it: dict, conf: dict) -> dict:
    brand_cat = it.get("brandCategory") or {}
    brand_kr = (it.get("brandName") or brand_cat.get("brandName") or "").strip()
    brand_en = (brand_cat.get("enName") or it.get("brandEnName") or "").strip()
    code = it.get("code", "")
    return {
        "sku": it.get("skuNo", ""),
        "brand_kr": brand_kr,
        "brand_en": brand_en,
        "category": _extract_category(it),
        "product_name": it.get("productNameForDisp") or it.get("name") or "",
        "ref_no": it.get("refNo", "") or code,
        "soldout": it.get("soldOutYn") == "Y",
        "detail_url": conf["detail_pc"].format(code=code) if code else "",
    }


def _lookup_by_skus(skus: List[str], site: str) -> dict:
    """SKU 목록을 해당 몰에서 일괄 조회 (세션·CSRF 토큰 재사용). 반환: {sku: item}"""
    if not skus:
        return {}
    conf = SITE_CONF[site]
    sess = _make_session(conf)
    token = _get_csrf(sess, conf, skus[0])
    out = {}
    for sku in skus:
        try:
            body = {
                "json": json.dumps({
                    "category": "", "size": "10", "page": 0,
                    "text": sku, "within": "", "query": sku,
                    "pagination": "", "condition": {"discountRate": "0"},
                }, ensure_ascii=False),
                "CSRFToken": token,
            }
            r = sess.post(conf["ajax"], data=body,
                          headers={"X-Requested-With": "XMLHttpRequest"}, timeout=15)
            r.raise_for_status()
            results = r.json().get("results", [])
            hit = next((it for it in results if it.get("skuNo") == sku), None)
            if hit:
                out[sku] = _item_from_hit(hit, conf)
        except Exception as e:
            logger.warning("SKU 일괄 조회 실패 site=%s sku=%s: %s", site, sku, e)
    return out


def _search_keyword(keyword: str, site: str = "kr", size: int = 50) -> list[dict]:
    conf = SITE_CONF[site]
    sess = _make_session(conf)
    token = _get_csrf(sess, conf, keyword)

    body = {
        "json": json.dumps({
            "category": "", "size": str(size), "page": 0,
            "text": keyword, "within": "", "query": keyword,
            "pagination": "", "condition": {"discountRate": "0"},
        }, ensure_ascii=False),
        "CSRFToken": token,
    }
    r = sess.post(conf["ajax"], data=body,
                  headers={"X-Requested-With": "XMLHttpRequest"}, timeout=15)
    r.raise_for_status()
    results = r.json().get("results", [])

    return [_item_from_hit(it, conf) for it in results]


@app.get("/api/search")
async def search_keyword(keyword: str, site: str = "kr"):
    _site_conf(site)
    keyword = keyword.strip()
    if not keyword:
        raise HTTPException(status_code=400, detail="검색어를 입력해주세요")
    if len(keyword) < 2:
        raise HTTPException(status_code=400, detail="검색어를 2자 이상 입력해주세요")

    loop = asyncio.get_running_loop()
    # 검색어는 항상 국문몰 기준으로 입력받는다 (외국몰도 국문 검색어 → 국문몰 검색으로 SKU 확보)
    kr_items = await loop.run_in_executor(None, _search_keyword, keyword, "kr")
    if site == "kr":
        return kr_items

    # 외국몰: 국문몰 검색으로 나온 SKU들을 해당 몰에서 일괄 조회해 결과값만 대체
    label = {"cn": "중문몰", "en": "영문몰", "jp": "일문몰"}[site]
    skus = [it["sku"] for it in kr_items if it["sku"]]
    n_workers = min(6, len(skus)) or 1
    chunks = [skus[i::n_workers] for i in range(n_workers) if skus[i::n_workers]]
    maps = await asyncio.gather(
        *[loop.run_in_executor(None, _lookup_by_skus, c, site) for c in chunks])
    site_map = {sku: item for m in maps for sku, item in m.items()}

    items = []
    for it in kr_items:
        found = site_map.get(it["sku"])
        if found:
            items.append(found)
        else:
            # 해당 몰 미판매 — 국문몰 정보로 표시하되 미판매 표기
            items.append({**it, "product_name": f"[{label} 미판매] {it['product_name']}",
                          "detail_url": "", "soldout": False})
    return items


@app.get("/api/lookup")
async def lookup(sku: str, site: str = "kr"):
    _site_conf(site)
    sku = sku.strip()
    if not sku:
        raise HTTPException(status_code=400, detail="SKU 번호를 입력해주세요")

    loop = asyncio.get_running_loop()
    result = await loop.run_in_executor(None, _fetch_product, sku, site)

    if not result:
        raise HTTPException(status_code=404, detail="상품을 찾을 수 없습니다. SKU 번호를 확인해주세요.")

    return result


class BatchRequest(BaseModel):
    skus: List[str]
    site: str = "kr"


@app.post("/api/batch")
async def batch_lookup(req: BatchRequest):
    _site_conf(req.site)
    skus = [s.strip() for s in req.skus if s.strip()]
    if not skus:
        raise HTTPException(status_code=400, detail="SKU 번호를 입력해주세요")
    if len(skus) > 50:
        raise HTTPException(status_code=400, detail="한 번에 최대 50개까지 조회할 수 있습니다")

    loop = asyncio.get_running_loop()
    sem = asyncio.Semaphore(5)  # 최대 5개 동시 요청

    async def fetch_one(sku: str):
        async with sem:
            try:
                result = await loop.run_in_executor(None, _fetch_product, sku, req.site)
                return result if result else {"sku": sku, "error": True}
            except Exception as e:
                logger.warning("배치 조회 실패 sku=%s: %s", sku, e)
                return {"sku": sku, "error": True}

    results = await asyncio.gather(*[fetch_one(sku) for sku in skus])
    return list(results)


class ExportRequest(BaseModel):
    kind: str  # "batch" | "keyword"
    keyword: str = ""
    site: str = "kr"
    rows: List[dict]


def _display_width(s: str) -> float:
    """한글 등 전각 문자는 폭 2배로 계산해 엑셀 열 너비 근사치를 구한다."""
    import unicodedata
    return sum(1.7 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in s)


def _autofit_columns(ws, min_width: float = 8, max_width: float = 70) -> None:
    for col_cells in ws.columns:
        max_len = 0.0
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, _display_width(str(cell.value)))
        ws.column_dimensions[col_cells[0].column_letter].width = max(min_width, min(max_len + 2, max_width))


def _xlsx_response(wb, filename: str) -> Response:
    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"export.xlsx\"; filename*=UTF-8''{quote(filename)}"
        },
    )


@app.post("/api/export")
async def export_excel(req: ExportRequest):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    _site_conf(req.site)

    wb = Workbook()
    ws = wb.active
    head_font = Font(bold=True, color="0B2E5C")
    link_font = Font(color="1F6FEB", underline="single")

    brand_col = "브랜드명" if req.site != "kr" else "국문 브랜드명"
    ws.title = "키워드검색" if req.kind == "keyword" else "SKU조회"
    headers = ["#", brand_col, "영문 브랜드명", "상품유형", "상품명", "SKU.NO", "REF.NO", "상품 페이지"]

    ws.append(headers)
    for c in ws[1]:
        c.font = head_font

    for i, r in enumerate(req.rows, start=1):
        if r.get("error"):
            ws.append([i, f"SKU {r.get('sku', '')} — 조회 실패"])
            continue
        ws.append([i, r.get("brand_kr", ""), r.get("brand_en", ""), r.get("category", ""),
                   r.get("product_name", ""), r.get("sku", ""), r.get("ref_no", ""), ""])
        link_col = len(headers)

        url = r.get("detail_url")
        cell = ws.cell(row=ws.max_row, column=link_col)
        if url:
            cell.value = "상품 페이지"
            cell.hyperlink = url
            cell.font = link_font
        else:
            cell.value = "—"

    _autofit_columns(ws)
    ws.freeze_panes = "A2"

    filename = f"키워드검색_{req.keyword}.xlsx" if req.kind == "keyword" else "SKU조회.xlsx"
    return _xlsx_response(wb, filename)
